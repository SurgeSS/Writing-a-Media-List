def M_write():
    import openpyxl

    dest_filename = input("What is the name of this new list? ")
    wb = openpyxl.load_workbook(dest_filename)

    #dest_filename = 'Media_List_Writer.xlsx'

    # Activate first sheet
    ws1 = wb.active

    # Set names for each of the sheets
    s1 = "Anime"
    s2 = "TV"
    s3 = "background data"

    # Create and title all sheets
    ws1.title = s1
    ws2 = wb.create_sheet(title=s2)
    ws3 = wb.create_sheet(title=s3)

    # Setup the first sheet
    ws1['A1'] = "Title"
    ws1['B1'] = "Status"

    # Setup the second sheet
    ws2['A1'] = "Title"
    ws2['B1'] = "Status"

    # Setup the third sheet
    ws3['A1'] = 1
    ws3['A2'] = 1

    # Save the document
    wb.save(filename = dest_filename)
